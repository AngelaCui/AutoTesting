# -*- coding: utf-8 -*-
"""
Created on Wed Jun 21 14:13:14 2017

@author: admin
"""

"""
说明：
通信协议解析程序

使用方法：
在main函数中‘修改串口号’处填写接收数据所使用串口编号，然后点击运行。
停止运行可按Crtl+C或点击console中'stop current command'按钮，即停止收数解析。

显示结果说明：
'Edge Output Error':边界信息解析失败
'CRC Check Error':CRC校验失败
'Edge Number Error':输出文件中‘边界编号’读取不成功
'Analysis Failed':解析失败
'Port Open Failed':串口调试失败
'Timing error/Exit':退出程序

'File Saved':已成功解析并保存
"""


import serial
import threading
import time
import math
from xlutils.copy import copy
from openpyxl import load_workbook

# 多线程方法

'''封装缓存区'''
class Buffer():
    def __init__(self,data,dataend = 0,flag = 0):
        self.data = data
        self.dataend = dataend
        self.flag = flag
        
'''创建新线程-串口读取数据并存储'''
class GetData(threading.Thread):
    def __init__(self, lock, threadName, ser, buffer):
        super(GetData, self).__init__(name = threadName)  
        self.lock = lock
        self.buffer = buffer
        self.ser = ser   
       
    def run(self):
        try:
            #self.buffer.data = [hex(ord(self.ser.read(1)))] 
            # 保留这句原因是在单独运行此函数情况下，必须先读入第一个数，ser.inWaiting才会正常工作，否则始终为0 但目前程序运行皆正常
            while True:
                if self.ser.inWaiting() != 0:
                    self.buffer.data.append(hex(ord(self.ser.read(1))))
                '''清理Buffer中无效数据部分'''
                if self.buffer.flag == 1:
                    self.lock.acquire()
                    try:
                        self.buffer.data = self.buffer.data[self.buffer.dataend+1:]
                        self.buffer.flag = 0
                    finally:
                        self.lock.release()
        except:
            return 'error'
    
'''确定帧头帧尾，提取数据帧'''
def GetDatapack(buffer):
    try:
        datahead = ''
        for i in range(len(buffer.data)):
            if buffer.data[i] == '0xff':
                if buffer.data[i+1] == '0xfb':
                    datahead = i
                    # datahead found         
                    for j in range(datahead +2, len(buffer.data)):
                        if buffer.data[j] == '0xff':
                            if buffer.data[j+1] == '0xfb':
                                datahead = j
                                # new datahead
                            elif buffer.data[j+1] == '0xfe':
                                buffer.dataend = j+1
                                datapack = buffer.data[datahead:buffer.dataend + 1]
                                # 检测到完整数据帧，提取为datapack
                                buffer.flag = 1
                                return datapack  
        else:
            return []
    except:
        return []
'''CRC校验'''
def CRC16(data):
    CRC16_XMODEM_TABLE = [
        0,4129, 8258, 12387, 16516, 20645, 24774, 28903, 33032, 37161, 41290, 45419, 49548, 53677,
	57806, 61935, 4657, 528, 12915, 8786, 21173, 17044, 29431, 25302, 37689, 33560, 45947, 41818,
	54205, 50076, 62463, 58334,9314, 13379, 1056, 5121, 25830, 29895, 17572, 21637, 42346, 46411,
	34088, 38153, 58862, 62927, 50604, 54669, 13907, 9842, 5649, 1584, 30423, 26358, 22165, 18100, 
	46939, 42874, 38681, 34616, 63455, 59390, 55197, 51132, 18628, 22757, 26758, 30887, 2112, 6241, 
	10242, 14371, 51660, 55789, 59790, 63919, 35144, 39273, 43274, 47403, 23285, 19156, 31415, 
	27286, 6769, 2640, 14899, 10770, 56317, 52188, 64447, 60318, 39801, 35672, 47931, 43802, 27814, 
	31879, 19684, 23749, 11298, 15363, 3168, 7233, 60846, 64911, 52716, 56781, 44330, 48395, 36200, 
	40265, 32407, 28342, 24277, 20212, 15891, 11826, 7761, 3696, 65439, 61374, 57309, 53244, 48923, 
	44858, 40793, 36728, 37256, 33193, 45514, 41451, 53516, 49453, 61774, 57711, 4224, 161, 12482, 
	8419, 20484, 16421, 28742, 24679, 33721, 37784, 41979, 46042, 49981, 54044, 58239, 62302, 689, 
	4752, 8947, 13010, 16949, 21012, 25207, 29270, 46570, 42443, 38312, 34185, 62830, 58703, 54572, 
	50445, 13538, 9411, 5280, 1153, 29798, 25671, 21540, 17413, 42971, 47098, 34713, 38840, 59231, 
	63358, 50973, 55100, 9939, 14066, 1681, 5808, 26199, 30326, 17941, 22068, 55628, 51565, 63758, 
	59695, 39368, 35305, 47498, 43435, 22596, 18533, 30726, 26663, 6336, 2273, 14466, 10403, 52093, 
	56156, 60223, 64286, 35833, 39896, 43963, 48026, 19061, 23124, 27191, 31254, 2801, 6864, 10931, 
	14994, 64814, 60687, 56684, 52557, 48554, 44427, 40424, 36297, 31782, 27655, 23652, 19525, 15522, 
	11395, 7392, 3265, 61215, 65342, 53085, 57212, 44955, 49082, 36825, 40952, 28183, 32310, 20053, 
	24180, 11923, 16050, 3793, 7920
        ] # CRC检查表
                
    us_crc = 0x0000
    us_count = 0
    us_cnt = 0
    
    for us_cnt in range(len(data)):     
        us_count = ((us_crc >> 8)^(int(data[us_cnt],16)&0xff))
        us_crc = ((us_crc << 8)&0xffff)^CRC16_XMODEM_TABLE[us_count]
        
    return us_crc
    
'''通信数据正文解析'''
def DataProcess(datapack):
    maindata = datapack[11:-4]
    for i in range(len(maindata)):
        maindata[i] = int(maindata[i],16)
        
    while True:
        if maindata[0] != 0xa2:
            break
        else:
            if maindata[5] >0 and maindata[5] <= 20:
                blocknum = maindata[5]
                '''闭塞分区状态'''
                stat = {
                    0b0001:"正常占用",
                    0b0010:"空闲",
                    0b0011:"故障占用",
                    0b0100:"失去分路",
                    0b0101:"出清（失去分路延时中）",
                    0b0110:"正常占用（越站调车）"}
                block_stat = [] 
                for i in range(int(blocknum/2)):
                    stat1 = stat[(maindata[6+i]>>4) & 0xF] # 字节内前一闭塞分区状态
                    stat2 = stat[maindata[6+i] & 0xF] # 字节内后一闭塞分区状态
                    block_stat.append(stat1)
                    block_stat.append(stat2)
                if blocknum % 2 == 1:
                    stat3 = stat[(maindata[6+int(blocknum/2)]>>4) & 0xF]
                    block_stat.append(stat3)
                '''闭塞分区状态结束时所处数据链中位置'''
                stat_end_num = 5 + math.ceil(blocknum/2)
                '''闭塞分区行车区间信息结束时位置'''
                info_end_num = stat_end_num + blocknum
            elif maindata[5] < 0 or maindata[5] > 20:
                return 'Exceeded block number'
            else:
                info_end_num = 5
            '''区间边界数量'''
            pos = info_end_num + 1
            edgenum = maindata[pos]
            
            edge_id = [] 
            edge_SA = []
            edge_block = []
            
            '''边界信号许可类型'''
            e_SA = { 
                0b00:"无信号许可",
                0b01:"发起信号许可",
                0b10:"应答信号许可",
                0b11:"故障（按无信号许可处理）"
                }
            '''边界闭塞分区状态  '''  
            e_block = { 
                0b001:"正常占用",
                0b010:"失去分路",
                0b011:"故障占用",
                0b100:"空闲"
                }
                
            try:
                for i in range(edgenum):
                    edge_id.append(maindata[pos+1]) 
                    edge_SA.append(e_SA[(maindata[pos+2]>>3)&0b11])
                    edge_block.append(e_block[(maindata[pos+2]>>5)&0b111])
                    pos += 3
            except:
                print('Edge Output Error')
            return {
                "闭塞分区数量":blocknum,"闭塞分区状态":block_stat,
                "区间边界数量":edgenum,"边界行车区间ID":edge_id,
                "边界信号许可类型":edge_SA,"边界闭塞分区状态":edge_block}

'''写入Excel单元格'''
def Output(result, file, pos):
    ws = file.get_sheet_by_name('维护终端信息')
    ws.cell(row = pos[0], column = pos[1], value = result)
    
'''收数后所有处理步骤'''
def MainProcess(ser,Data_get, Tmp):
    try:
        '''提取完整数据帧'''
        Tmp = GetDatapack(Data_get)
        if Tmp != []:
            Datapack = Tmp
            Tmp = []

            '''校验'''
            CRCCode0 = Datapack[-4:-2]
            CRCCode = int(CRCCode0[0],16)*256 + int(CRCCode0[1],16)
            CRCData = Datapack[:11]
            CRCCheck = CRC16(CRCData)
            if CRCCode != CRCCheck:
                print ('CRC Check Error')
                                
            '''通信协议解析'''
            Result = DataProcess(Datapack)
            
            '''确定信息输出文件名称'''
            '''确定测试用例执行步骤序号'''
            wb = load_workbook('QJK自动化测试用例.xlsx')
            ws = wb.get_sheet_by_name('接车口进站正常行车用例')
            ws1 = wb.get_sheet_by_name('维护终端信息')
            step = ws.cell('C1').value
            try:
                k = ws1.cell('F1').value -1
            except:
                print('Edge Number Error')
                k = 0
                
            '''边界信息输出至Excel'''
            Output(Result["边界行车区间ID"][k],wb,[step+3,2])
            Output(Result["边界信号许可类型"][k],wb,[step+3,3])
            Output(Result["边界闭塞分区状态"][k],wb,[step+3,4])
            
            '''闭塞分区信息输出至Excel'''
            blocknum = Result["闭塞分区数量"]
            Output(blocknum,wb,[1,9])
            for i in range(blocknum):
                Output(Result["闭塞分区状态"][i],wb,[step+3,i+5])
            
            '''保存结果'''
            Savename = "QJK自动化测试用例.xlsx"
            wb.save(Savename)
            print('File Saved')
        else:
            pass
    except:
        #ser.close()
        print('Analysis Failed')
    
'''定时循环'''
def ATimer(interval, ser, Data_get, Tmp): 
    time_remaining = interval - time.time()%interval
    time.sleep(time_remaining)
    MainProcess(ser, Data_get, Tmp)

'''主函数'''
def main():
    '''初始化各项'''
    lock = threading.Lock()
    Data_get = Buffer([])
    Tmp = []
    
    '''打开串口'''
    '''在此处修改串口号''' #修改串口号
    portnum = '1'
    try:
        ser = serial.Serial(
            port='COM'+portnum,
            baudrate=115200,
            parity=serial.PARITY_ODD,
            stopbits=serial.STOPBITS_ONE,
            bytesize=serial.EIGHTBITS
            )
    except:
        print('Port Open Failed')
    '''运行GetData线程'''
    GetData(lock, 'GetData', ser, Data_get).start()
    '''定时器-0.12s'''
    '''定时循环Mainprocess'''
    while True:
        try:
            ATimer(0.1, ser, Data_get, Tmp)
            #MainProcess(ser, Data_get, Tmp)
            #t = threading.Timer(0.12,MainProcess,[ser, Data_get, Tmp])
            #t.start()
            '''第一行为自设定时器，定时运行'''# 试验效果好
            '''第二行为无定时器，始终运行'''# 试验效果好
            '''第三四行为自带定时器，另起线程定时运行'''# 试验效果最差
        except: 
            print('Timing error')
            break
        
    ser.close()
    print('Exit')
    
if __name__ == '__main__':
    main()