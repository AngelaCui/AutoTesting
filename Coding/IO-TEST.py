# -*- coding: UTF-8 -*-
import xlrd
import xlwt
import wx
import docx 
import os
import win32com
from docx.api import Document
import wx
from audioop import add
from _ast import Pass
from pip._vendor.colorama.ansi import Style
from xlrd import open_workbook
import openpyxl
from _codecs import decode

""" 读取EXCEL表格"""
def ReadExcelCell(bk,shName,Row,Col ):
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    return sh.row_values(Row)[Col]

""" 获取站ID号"""
def GetStationId(bk):
    shName="车站信息".decode("utf8")
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    return sh.row_values(4)[0]

""" 获取站ID号"""
def GetStationName(bk):
    shName="车站信息".decode("utf8")
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    return sh.row_values(4)[1]
""" 获取区间名称，区段名"""
def GetZoneInfor(bk,listName,listQD,statId):
    shName="发车口信息".decode("utf8")
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    for i in range(4,sh.nrows):
        s=sh.row_values(i)[2]
        if s==statId:
            listName.append(sh.row_values(i)[1])
            listQD.append(sh.row_values(i)[12])
""" 获取区间名称，区段名"""
def GetTrackSectName(bk,ListTrackSectId,stationId,ListTrackSectName):
    shName="闭塞分区信息".decode("utf8")
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    nrows = int(sh.row_values(1)[0])
    for trackId in ListTrackSectId:
        for i in range(0,nrows):
            s= int(sh.row_values(i+4)[0])
            s1 = '%d'%s
            if s1== trackId:
                ListTrackSectName.append(sh.row_values(i +4)[1])  
""" 获取闭塞分区名称"""
def MakeTrackDic(station_bk,IO_bk,dic):
    list_Track=[]
    lst_FC=[]
    list_dic = {}
    stat_id = GetStationId(station_bk)
    GetZoneInfor(station_bk,lst_FC,list_Track,stat_id)
    i = 0
    for ZoneName in lst_FC:
        Track = list_Track[i].split('\\')
        listTrackName = []
        GetTrackSectName(station_bk,Track,stat_id,listTrackName)
        dic[ZoneName] = listTrackName
        i = i +1
""" 获取车站名称"""                   
def  GetComStaName(bk,list):
    shName="通信车站表".decode("utf8")
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    listName = []
    
    listName = sh.col_values(1)   
    for i in listName:
        list.append(i)        
""" 获取解锁盘名称""" 
def  GetJSPStaName(bk,list):
    shName="解锁盘配置表".decode("utf8")
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    listName = []
    listName = sh.col_values(1)   
    for i in listName:
        list.append(i)        
             
"""读取码位表的继电器采集信息""" 
def  GetINIO(bk,dic):  
    shName="继电器采集表".decode("utf8")
    i = 0
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    for i in range(1,6): 
        if i < sh.ncols:
            dic[i] = sh.col_values(i)
    """for i in range(1,sh.nrows):
     list.append(sh.row_values(i)[DINum])"""
    return i
"""读取码位表的继电器采集信息""" 
def  GetJSPButtonInfor(bk,JSPNameList,List):  
    dic = {}
    shName="解锁盘按钮名称".decode("utf8")
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    
    dic[0] = sh.col_values(1)
    shName="解锁盘指示灯名称".decode("utf8")
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    dic[1] = sh.col_values(1)   
    """for i in range(1,sh.nrows):
     list.append(sh.row_values(i)[DINum])"""
    for name in JSPNameList:
        
        for JSPA in dic[0]:
            if name[:-3] in JSPA:
                List.append(JSPA)
        for JSPD in dic[1]:
            if name[:-3] in JSPD:
                List.append(JSPD)
           
"""读取码位表的继电器驱动信息""" 
def  GetOUTIO(bk,dic):  
    shName="继电器驱动表".decode("utf8")
    i = 0
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    for i in range(1,6): 
        if i < sh.ncols:
            dic[i] = sh.col_values(i)
    return i
"""读取轨道区段信息""" 
def GetTrackInfor(bk,dic):
    shName="轨道区段信息".decode("utf8")
    try:  
        sh = bk.sheet_by_name(shName) 
    except:  
        print "no sheet in %s named " %shName #
    list = sh.col_values(1)
    dic[0] = list
"""写闭塞分区检查表"""     
def OpenWriteExcel(rd,dic):
    ws= rd.get_sheet_by_name("闭塞分区检查表".decode('utf8'))
    Row = 0
    for num in dic:
        list = []
        list = dic[num]
        Row = Row +1
        num = 2 + Row
        for i in range(1,6):
            ws.cell(row =num,column = (i + 1),value = list[i-1])
        for i in range(1,4):
            ws.cell(row =num,column = (i + 19),value = list[4+i])
        for i in range(1,4):
            ws.cell(row =num,column = (i + 23),value = list[7+i])
    return
"""根据闭塞分区名查找轨道区段名 """   
def  findTrackName(Zonedic,TrackDic,ZoneTrackDic):
    i = 0
    for zoneName in Zonedic:
        ListTrack = []
        ListTrack = Zonedic[zoneName]
        findList = TrackDic[0]
        for Track in ListTrack:
            i=i+1
            list = []
            list.append(Track)
            list.append(zoneName)
            for name in findList[4:]:
                if Track[:-1] in name:
                    if Track[:2] ==name[:2]:
                        list.append(name)  
            ZoneTrackDic[i] = list
"""写区间平面图检查表 """ 
def writePlaneGraph(rd,dic):
     ws= rd.get_sheet_by_name("区间平面图检查表".decode('utf8')) 
     i=0
     for Num in dic:
         list = []
         i = i+1
         j = 0
         list = dic[Num]
         for Name in list:
             j = j+1
             ws.cell(row =1 +i,column = j+1,value = Name)
             
"""写通信状态检查表"""      
def writeComTableExcel(rd,list,type):
    ws= rd.get_sheet_by_name("通信接口检查表".decode('utf8'))
    i=0
    if type == 1:
        col = 1
    else:
        col = 4
    for staname in list[1:]:
        ws.cell(row =4 + 2*i,column = col,value = staname)
        i = i+1
"""写封面名称"""  
def writeStaNameTableExcel(rd,staname):
    ws= rd.get_sheet_by_name("封面".decode('utf8'))
    ws.cell(row =6,column = 3,value = staname)
def writeStopFunTable(rd,list):
    ws = rd.get_sheet_by_name("功能停用检查表".decode('utf8'))
    j = 0
    for n in list :
        ws.cell(row =4+j,column = 1,value = n)
        j = j +1

"""写3列信息表操作"""  
def write3dotTable(rd,dic,Listname):
    ws= rd.get_sheet_by_name(Listname)
    i = 0
    for item in dic:
        list = []
        list = dic[item]
        j = 0
        for n in list :
            ws.cell(row =2+i,column = 1+j,value = n)
            j = j +1
        i = i +1
  
def findInfor(name, dic,dotNum):
    ret = 0
    for board in dic:
        list = dic[board]
        i=0;
        for n in list:
            if name in n:
                ret = 1
                break
            i = i +1
        if ret == 1:
            break
    if ret != 1:
       board = 0
       i = 0 
    dotNum.append(board)
    dotNum.append(i) 
   
def  findStrInDic(listname,dic,DicOut):
    ret = 0
    cnt = 0;
    for name in listname: 
        for Board in dic :
            list= []
            list = dic[Board]
            i=0
            for n in list:
                if name in n:
                    listName = []
                    listName.append(n)
                    listName.append(Board)
                    listName.append(i)  
                    DicOut[cnt] = listName
                    cnt = cnt + 1
                i=i+1
"""界面显示"""            
class MainFrame(wx.Frame):
    def __init__(self):
        wx.Frame.__init__(self,None,-1,"用例生成工具".decode("utf8"),size = (1000,300))
        panel = wx.Panel(self,-1)
        butnchstationFile = wx.Button(panel,-1,"选择车站文件".decode("utf8"),pos=(50,10))
        butnchJCCodeTable = wx.Button(panel,-1,"选择码位表".decode("utf8"))
        butnchTestFile = wx.Button(panel,-1,"选择测试用例模板".decode("utf8"))
        butnRun = wx.Button(panel,-1,"运行".decode("utf8"))
        sizer1 = wx.BoxSizer(wx.HORIZONTAL)
        self.FileStation = wx.TextCtrl(panel,wx.NewId(),"",size=(160,40))
        sizer1.Add(butnchstationFile,2,wx.EXPAND|wx.ALL,5)
        sizer1.Add(self.FileStation,2,wx.EXPAND|wx.ALL,5)
        self.FileCode = wx.TextCtrl(panel,wx.NewId(),"",size=(160,40))
        sizer2 = wx.BoxSizer(wx.HORIZONTAL)
        sizer2.Add(butnchJCCodeTable,2,wx.EXPAND|wx.ALL,5)
        sizer2.Add(self.FileCode,2,wx.EXPAND|wx.ALL,5)
       
        self.FileTest = wx.TextCtrl(panel,wx.NewId(),"",size=(160,40))
        sizer3 = wx.BoxSizer(wx.HORIZONTAL)
        sizer3.Add(butnchTestFile,2,wx.EXPAND|wx.ALL,5)
        sizer3.Add(self.FileTest,2,wx.EXPAND|wx.ALL,5)
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(sizer1,2,wx.EXPAND|wx.ALL,5)
        sizer.Add(sizer2,2,wx.EXPAND|wx.ALL,5)
        sizer.Add(sizer3,2,wx.EXPAND|wx.ALL,5)
        sizer.Add(butnRun,2,wx.EXPAND|wx.ALL,5)
        self.StationFile = " "
        self.CodeFile = " "
        self.TestFile = " "
        panel.SetSizer(sizer)
        self.Bind(wx.EVT_BUTTON, self.OnButnRun, butnRun)
        self.Bind(wx.EVT_BUTTON, self.OnbutnchstationFile, butnchstationFile)
        self.Bind(wx.EVT_BUTTON, self.OnbutnchJCCodeTable, butnchJCCodeTable)
        self.Bind(wx.EVT_BUTTON, self.OnbutnchTestFile, butnchTestFile)
        
    def OnButnRun(self,event):
        if self.FileStation != " " :
            if self.CodeFile != " ":
                if self.TestFile != " ":
                    self.exeHandle()
    def OnbutnchstationFile(self,event):
        dialog = wx.FileDialog(self,"open file station",os.getcwd(), style=wx.OPEN,wildcard="*.xlsx")  
        if dialog.ShowModal()==wx.ID_OK:
            self.StationFile = dialog.GetPath()
            self.FileStation.SetValue(self.StationFile)
            dialog.Destroy()
        
    def OnbutnchJCCodeTable(self,event):
        dialog = wx.FileDialog(self,"open file code",os.getcwd(), style=wx.OPEN,wildcard="*.xlsx")  
        if dialog.ShowModal()==wx.ID_OK:
            self.CodeFile = dialog.GetPath()
            self.FileCode.SetValue(self.CodeFile)
            dialog.Destroy()  
            
    def OnbutnchTestFile(self,event):
        dialog = wx.FileDialog(self,"Open file 测试用例模板",os.getcwd(), style=wx.OPEN,wildcard="*.xlsx")  
        if dialog.ShowModal()==wx.ID_OK:
            self.TestFile = dialog.GetPath()
            self.FileTest.SetValue(self.TestFile)
            dialog.Destroy()  
            
    def  exeHandle(self):
        """打开各文件"""
        rdtest = openpyxl.load_workbook(self.TestFile)
        bkFile = xlrd.open_workbook(r"%s"%self.StationFile)
        bkCode = xlrd.open_workbook(r"%s"%self.CodeFile)
        """获取闭塞分区信息"""
        Zonedic = {} 
        MakeTrackDic(bkFile,bkCode,Zonedic)
        
        """获取采集表信息"""
        IOIN_Dic = {}
        GetINIO(bkCode,IOIN_Dic)
        """获取驱动表信息"""
        IOOUT_Dic = {}
        GetOUTIO(bkCode,IOOUT_Dic)
        TrackList = {}
        """获取轨道区段表信息"""
        GetTrackInfor(bkFile,TrackList)
        
        """在采集信息表中搜索各发车口FJ信息,写入方向初始化检查表"""
        Track_Dic = {}
        dic = {}
        listfind = ["FJ"]
        findStrInDic(listfind,IOIN_Dic,dic)
        ListName ="方向初始化检查表".decode("utf8")
        write3dotTable(rdtest,dic,ListName)
        """在采集信息表中搜索各发车口LXJ|YXJ|FSJ信息,写入进出站检查表"""
        listfind = ["LXJ","YXJ","FSJ"]
        dic = {}
        findStrInDic(listfind,IOIN_Dic,dic)
        ListName = "进出站检查表".decode("utf8")
        write3dotTable(rdtest,dic,ListName)
        """根据闭塞分区信息，搜索采集表和驱动表，获取点位信息,写入闭塞分区表"""
        i=0
        ZoneTrackDic = {}
        findTrackName(Zonedic,TrackList,ZoneTrackDic)
        writePlaneGraph(rdtest,ZoneTrackDic)
        """闭塞分区表"""
        for zoneName in Zonedic:
            ListTrack = {}
            ListTrack = Zonedic[zoneName]
            for Track in ListTrack:
                i=i+1
                list = []
                list.append(Track)
                GJ_NAME = "QGJ-"+Track
                dotnum = []
                findInfor(GJ_NAME,IOIN_Dic,dotnum)
                if dotnum[0] != 0:
                    list.append(GJ_NAME)
                    list.append(dotnum[0])
                    list.append(dotnum[1])
                    list.append(zoneName)
                else:
                    dotnum = []
                    GJ_NAME = "GJ-"+Track
                    findInfor(GJ_NAME,IOIN_Dic,dotnum)
                    if dotnum[0] != 0:
                        list.append(GJ_NAME)
                    else:
                        GJ_NAME = "NO"
                        list.append(GJ_NAME)
                    list.append(dotnum[0])
                    list.append(dotnum[1])
                    list.append(zoneName)
                
                FHJ_NAME = Track+"-FHJ"
                
                dotnum = []
                findInfor(FHJ_NAME,IOIN_Dic,dotnum)
                if dotnum[0] != 0:
                    FHJ_NAME = Track+"-FHJ"
                    list.append(FHJ_NAME)
                    list.append(dotnum[0])
                    list.append(dotnum[1])
                else:
                    FHJ_NAME = "NO "
                    list.append(FHJ_NAME)
                    list.append(dotnum[0])
                    list.append(dotnum[1])
                
                FHJ_NAME = Track+"-FHJ"
                dotnum = []
                findInfor(FHJ_NAME,IOOUT_Dic,dotnum)
                if dotnum[0] != 0:
                    FHJ_NAME = Track+"-FHJ"
                    list.append(FHJ_NAME)
                    list.append(dotnum[0])
                    list.append(dotnum[1])
                else:
                    FHJ_NAME = "NO"
                    list.append(FHJ_NAME)
                    list.append(dotnum[0])
                    list.append(dotnum[1])
                Track_Dic[i] = list
        OpenWriteExcel(rdtest,Track_Dic)
        """获取通信站信息写入通信检查表"""
        listStaName = []
        GetComStaName(bkCode,listStaName)
        writeComTableExcel(rdtest,listStaName,1)
        """获取解锁盘信息写入通信检查表"""
        listStaName = []
        GetJSPStaName(bkCode,listStaName)
        writeComTableExcel(rdtest,listStaName,2)
        """获取"""
        listJSP = []
        GetJSPButtonInfor(bkCode,listStaName,listJSP)

        writeStopFunTable(rdtest,listJSP)
        stationName = GetStationName(bkFile)
        writeStaNameTableExcel(rdtest,stationName)
        SaveName = "C:\ConfRecorde\QJK-JS区间综合监控工程数据IO对点表".decode('utf8')+"(%s)"%stationName+".xlsx"
        print SaveName
        rdtest.save(SaveName)
if __name__ == '__main__':
    app = wx.PySimpleApp()
    frame = MainFrame()
    frame.Show()
    app.MainLoop()
