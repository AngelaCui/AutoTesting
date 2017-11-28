# -*- coding: utf-8 -*-
"""
Created on Fri Jun 16 15:46:20 2017

@author: admin
"""

# output into excel
import xlwt
import xlrd
from xlutils.copy import copy
import openpyxl
import re

#from data import dataprocess
blocknum = 5
block_stat = ['空闲','正常占用','故障占用','失去分路','出清']
block_info = [2,3,5,0,4]
edge_id = 3
edge_SA = ['应答信号许可']
edge_block = ['故障占用']
result = {"闭塞分区数量":blocknum,"闭塞分区状态":block_stat,
          "闭塞分区行车区间信息":block_info,"边界行车区间信息":edge_id,
          "边界信号许可类型":edge_SA,"边界闭塞分区状态":edge_block}
def read_excel():
    workbook = xlrd.open_workbook('QJK自动化测试用例.xlsx')
    
    wb = copy(workbook)
    
    ws = wb.get_sheet(0)
    wss = wb.get_sheet_by_name('维护终端状态')
    #ws.write(0,8,blocknum)
    ws.cell(row = 0,column = 8,value = blocknum)
    #ws.write(3,1,result["边界行车区间信息"])
    #ws.write(3,2,result["边界信号许可类型"])
    #ws.write(3,3,result["边界闭塞分区状态"])
    for i in range(blocknum):
        tmp = result["闭塞分区状态"][i]
        #ws.write(3,4+i,tmp)
    
        
    
    wb.save('QJK自动化测试用例.xlsx')
        
read_excel()


import xlrd
from xlutils.copy import copy

def read_excel(sheet_index):
    workbook = xlrd.open_workbook('QJK自动化测试用例.xlsx')
    sh = workbook.sheet_by_name('维护终端信息')
    
    wb = copy(workbook)
    ws = wb.get_sheet(sheet_index)
    dd = workbook.get_sheet_by_name('维护终端信息')
    b = dd.cell(row = 0,column = 0,value = 3)
    a = sh.cell(0,0)
        
    wb.save('QJK自动化测试用例.xlsx')
    return [a,b]
    
import openpyxl #import load_workbook
wb = load_workbook('QJK自动化测试用例.xlsx')
ws = wb.get_sheet_by_name('维护终端信息')
c = ws.cell(row = 3,column = 3).value
print('c = ',c)
ws.cell(row = 3, column = 3, value = 1)
d = ws.cell(row = 3,column = 3).value
print('d = ',d)