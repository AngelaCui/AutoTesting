# -*- coding: utf-8 -*-
"""
Created on Tue Jun 13 09:27:11 2017

@author: admin
"""

import serial
import math
import xlwt
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
import time



'''从串口读取数据并缓存''' # 只读其中一次发送的数据

'''添加：删除buffer中无效部分'''


'''确定帧头帧尾，提取完整数据帧'''
def GetDatapack(data): 
    try:
        for i in range(len(data)):
            if data[i] == '0xff':
                if data[i+1] == '0xfb':
                    datahead = i
                    break
        for j in range(datahead+13, len(data)):
            if data[j] == '0xff':
                if data[j+1] == '0xfe':
                    dataend = j+1
                    break
        return [dataend,data[datahead:dataend+1]]
    except:
        return []

'''提取数据帧后清空缓存区'''
def BufferDelete(data,dataend):
    data = data[dataend+1:]
    return data

'''计时器'''
def Timer_GetData(interv):
    while True:
        try:
            time_remaining = interv - time.time()%interv
            time.sleep(time_remaining)
            GetData
        except:
            return 'error'

'''CRC校验'''
def CRC16(data,crc = 0):
    CRC16_XMODEM_TABLE = [
        0x0000, 0x1021, 0x2042, 0x3063, 0x4084, 0x50a5, 0x60c6, 0x70e7,
        0x8108, 0x9129, 0xa14a, 0xb16b, 0xc18c, 0xd1ad, 0xe1ce, 0xf1ef,
        0x1231, 0x0210, 0x3273, 0x2252, 0x52b5, 0x4294, 0x72f7, 0x62d6,
        0x9339, 0x8318, 0xb37b, 0xa35a, 0xd3bd, 0xc39c, 0xf3ff, 0xe3de,
        0x2462, 0x3443, 0x0420, 0x1401, 0x64e6, 0x74c7, 0x44a4, 0x5485,
        0xa56a, 0xb54b, 0x8528, 0x9509, 0xe5ee, 0xf5cf, 0xc5ac, 0xd58d,
        0x3653, 0x2672, 0x1611, 0x0630, 0x76d7, 0x66f6, 0x5695, 0x46b4,
        0xb75b, 0xa77a, 0x9719, 0x8738, 0xf7df, 0xe7fe, 0xd79d, 0xc7bc,
        0x48c4, 0x58e5, 0x6886, 0x78a7, 0x0840, 0x1861, 0x2802, 0x3823,
        0xc9cc, 0xd9ed, 0xe98e, 0xf9af, 0x8948, 0x9969, 0xa90a, 0xb92b,
        0x5af5, 0x4ad4, 0x7ab7, 0x6a96, 0x1a71, 0x0a50, 0x3a33, 0x2a12,
        0xdbfd, 0xcbdc, 0xfbbf, 0xeb9e, 0x9b79, 0x8b58, 0xbb3b, 0xab1a,
        0x6ca6, 0x7c87, 0x4ce4, 0x5cc5, 0x2c22, 0x3c03, 0x0c60, 0x1c41,
        0xedae, 0xfd8f, 0xcdec, 0xddcd, 0xad2a, 0xbd0b, 0x8d68, 0x9d49,
        0x7e97, 0x6eb6, 0x5ed5, 0x4ef4, 0x3e13, 0x2e32, 0x1e51, 0x0e70,
        0xff9f, 0xefbe, 0xdfdd, 0xcffc, 0xbf1b, 0xaf3a, 0x9f59, 0x8f78,
        0x9188, 0x81a9, 0xb1ca, 0xa1eb, 0xd10c, 0xc12d, 0xf14e, 0xe16f,
        0x1080, 0x00a1, 0x30c2, 0x20e3, 0x5004, 0x4025, 0x7046, 0x6067,
        0x83b9, 0x9398, 0xa3fb, 0xb3da, 0xc33d, 0xd31c, 0xe37f, 0xf35e,
        0x02b1, 0x1290, 0x22f3, 0x32d2, 0x4235, 0x5214, 0x6277, 0x7256,
        0xb5ea, 0xa5cb, 0x95a8, 0x8589, 0xf56e, 0xe54f, 0xd52c, 0xc50d,
        0x34e2, 0x24c3, 0x14a0, 0x0481, 0x7466, 0x6447, 0x5424, 0x4405,
        0xa7db, 0xb7fa, 0x8799, 0x97b8, 0xe75f, 0xf77e, 0xc71d, 0xd73c,
        0x26d3, 0x36f2, 0x0691, 0x16b0, 0x6657, 0x7676, 0x4615, 0x5634,
        0xd94c, 0xc96d, 0xf90e, 0xe92f, 0x99c8, 0x89e9, 0xb98a, 0xa9ab,
        0x5844, 0x4865, 0x7806, 0x6827, 0x18c0, 0x08e1, 0x3882, 0x28a3,
        0xcb7d, 0xdb5c, 0xeb3f, 0xfb1e, 0x8bf9, 0x9bd8, 0xabbb, 0xbb9a,
        0x4a75, 0x5a54, 0x6a37, 0x7a16, 0x0af1, 0x1ad0, 0x2ab3, 0x3a92,
        0xfd2e, 0xed0f, 0xdd6c, 0xcd4d, 0xbdaa, 0xad8b, 0x9de8, 0x8dc9,
        0x7c26, 0x6c07, 0x5c64, 0x4c45, 0x3ca2, 0x2c83, 0x1ce0, 0x0cc1,
        0xef1f, 0xff3e, 0xcf5d, 0xdf7c, 0xaf9b, 0xbfba, 0x8fd9, 0x9ff8,
        0x6e17, 0x7e36, 0x4e55, 0x5e74, 0x2e93, 0x3eb2, 0x0ed1, 0x1ef0
        ]
    for byte in data:
        crc = ((crc << 8)&0xff00) ^ CRC16_XMODEM_TABLE[((crc >> 8)&0xff)^int(byte,16)]
    return crc & 0xffff

'''通信数据正文解析'''
def DataProcess(datapack):
    maindata = datapack[11:-4]
    while True:
        if maindata[0] != 0xa2:
            return 'Incorrect package'
        else:
            if maindata[5] >0 and maindata[5] <= 20:
                blocknum = maindata[5]
                stat = {
                    0b0001:"正常占用",
                    0b0010:"空闲",
                    0b0011:"故障占用",
                    0b0100:"失去分路",
                    0b0101:"出清（失去分路延时中）",
                    0b0110:"正常占用（越站调车）"}
                block_stat = [] # 依次为闭塞分区状态
                for i in range(int(blocknum/2)):
                    stat1 = stat[(maindata[6+i]>>4) & 0xF] # 字节内前一闭塞分区状态
                    stat2 = stat[maindata[6+i] & 0xF] # 字节内后一闭塞分区状态
                    block_stat.append(stat1)
                    block_stat.append(stat2)
                if blocknum % 2 == 1:
                    stat3 = stat[(maindata[6+int(blocknum/2)]>>4) & 0xF]
                    block_stat.append(stat3)
                # 闭塞分区状态结束时所处数据链中位置
                stat_end_num = 5 + math.ceil(blocknum/2)
                # 闭塞分区行车区间信息结束时位置
                info_end_num = stat_end_num + blocknum
            elif maindata[5] < 0 or maindata[5] > 20:
                return 'Exceeded block number'
            else:
                info_end_num = 5
            # 区间边界数量
            pos = info_end_num + 1
            edgenum = maindata[pos]
            
            edge_id = [] #行车区间ID
            edge_SA = []
            edge_block = []
            
            e_SA = { # 边界信号许可类型
                0b00:"无信号许可",
                0b01:"发起信号许可",
                0b10:"应答信号许可",
                0b11:"故障（按无信号许可处理）"
                }
            e_block = { # 边界闭塞分区状态
                0b001:"正常占用",
                0b010:"失去分路",
                0b011:"故障占用",
                0b100:"空闲"
                }
                
            for i in range(edgenum):
                edge_id.append(maindata[pos+1]) # 边界1 id
                edge_SA.append(e_SA[(maindata[pos+2]>>3)&0b11])
                edge_block.append(e_block[(maindata[pos+2]>>5)&0b111])
                pos += 3
            
            return {
                "闭塞分区数量":blocknum,"闭塞分区状态":block_stat,
                "区间边界数量":edgenum,"边界行车区间ID":edge_id,
                "边界信号许可类型":edge_SA,"边界闭塞分区状态":edge_block}

'''结果写入Excel对应sheet的单元格中'''
def Output(result, file, pos):
    ws = file.get_sheet_by_name('维护终端状态')
    ws.cell(row = pos[0], column = pos[1], value = result)

class DataBuffer():
    def __init__(self,data):
        self.data = data

'''主函数'''
def main():
    Buffer = DataBuffer()
    Tmp = []
    '''开启串口'''
    ser = serial.Serial( # 正式代码中放在主程序开头
            port = 'COM2',
            baudrate=115200,
            parity=serial.PARITY_ODD,
            stopbits=serial.STOPBITS_ONE,
            bytesize=serial.EIGHTBITS,
            timeout = 0.12
            )
            
    '''Timer循环范围开始''' #!!!!!!!!!!!!!!!!!!!!!!!!!!!!     
    '''绑定Timer，每0.12s开始GetData'''
    # 此处缺少Timer!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    #Buffer = GetData(ser,Buffer)
    GetData(ser,Buffer.data)
    
    '''定时提取完整数据帧'''
    Tmp = GetDatapack(Buffer.data)
    if Tmp != []:
        Buffer.dataend = Tmp[0]
        flag = 1  '''改成class'''
        datapack = Tmp[1]
        Tmp = []
    '''删除缓冲区内无效部分'''
    Buffer.data = Buffer.data[dataend+1:]
    
    '''Timer循环范围结束''' #!!!!!!!!!!!!!!!!!!!!!!!!!!!!     
    
    
    '''校验数据'''
    CRCCode = datapack[-4:-2]
    CRCData = datapack[:11]
    CRCCheck = CRC16(CRCData)
    if CRCCode != CRCCheck:
        return 'CRC Check Error'
        
    '''通信协议解析'''
    Result = DataProcess(datapack)
    
    '''确定测试用例执行步骤序号'''
    wb = load_workbook('QJK自动化测试用例.xlsx')
    ws = wb.get_sheet_by_name('接车口进站正常行车用例')
    step = ws.cell('C1').value
    
    '''边界信息输出至Excel'''
    for k in range(Result["区间边界数量"]):
        Output(Result["边界行车区间ID"][k],wb,[step+3,1])
        Output(Result["边界信号许可类型"][k],wb,[step+3,2])
        Output(Result["边界闭塞分区状态"][k],wb,[step+3,3])
        
    '''闭塞分区信息输出至Excel'''
    blocknum = Result["闭塞分区数量"]
    Output(blocknum,wb,[0,8])
    for i in range(blocknum):
        Output(Result["闭塞分区状态"][i],wb,[step+3,i+4])
        
    '''保存结果'''
    Savename = "QJK自动化测试用例.xlsx"
    wb.save(Savename)
    
    
if __name__ == '__main__':
    main()